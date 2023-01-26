import openpyxl


def row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_row


total_row_count = row_count("F:\\SQA\\Batch09\\AutomationBITM09\\File\\data.xlsx", "Sheet1")
print(total_row_count)


def column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_column


total_column_count = column_count("F:\\SQA\\Batch09\\AutomationBITM09\\File\\data.xlsx", "Sheet1")
print(total_column_count)


def read_data(file, sheet, reading_row, reading_column):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.cell(row=reading_row, column=reading_column).value

"""
data = read_data("F:\\SQA\\Batch09\\AutomationBITM09\\File\\data.xlsx", "Sheet1", 5, 1)
print(data)
"""


for r in range(1, total_row_count + 1):
    for c in range(1, total_column_count+1):
        data = read_data("F:\\SQA\\Batch09\\AutomationBITM09\\File\\data.xlsx", "Sheet1",r, c)
        print(data)

